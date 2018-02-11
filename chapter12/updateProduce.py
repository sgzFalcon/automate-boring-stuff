import openpyxl as xls
wb = xls.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']
PRICE_UPDATES = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

for row in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row, 1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row,2).value = PRICE_UPDATES[produceName]
wb.save('updatedProduceSales.xlsx')
